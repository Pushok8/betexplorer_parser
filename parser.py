import os
import datetime
from typing import Any
from urllib.parse import urlencode
from random import choice

import requests
import openpyxl
from requests import Response
from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.colors import Color

import make_pattern_xlsx
from annotations import url_type, column_name, numeric_str, name


HOST = "https://www.betexplorer.com"


def get_response_from_url(url: url_type, **request_parameters) -> Response:
    """
    Get response by url with parameters(if they are have).

    :param url(url_type) -> Url without parameters.
    :param **request_parameters -> Parameters specified in requests.get method.

    :return(Response) -> Response from url with parameters.
    """
    user_agent: str = choice(open('user_agents.txt').readlines()).strip()
    request_parameters['headers'] = request_parameters.get('headers', {}) | {'User-Agent': user_agent}
    response: Response = requests.get(url, **request_parameters)

    return response


def get_list_of_links_to_matches(url: url_type) -> list[url_type]:
    """
    Whether the user gets a date interval and parses links in the range of specified dates for each day.

    :param url

    :return list_of_links_to_matches(list[url_type])
    """
    list_of_links_to_matches: list[url_type] = []

    date_start: datetime.date = datetime.date.fromisoformat(input('Введите дату, с которой нужно начать'
                                                                  '(гггг-мм-дд): '))
    date_end: datetime.date = datetime.date.fromisoformat(input(f'Введите дату, на которой нужно закончить'
                                                                f'(если не указать, будет '
                                                                f'{datetime.date.today()}): '))

    if date_end == '':
        date_end = datetime.date.today()

    while date_start != date_end:
        date_for_url = {'year': str(date_start).split('-')[0],
                        'month': str(date_start).split('-')[1],
                        'day': str(date_start).split('-')[2]}
        url += f'?{urlencode(date_for_url)}'
        page_with_matches: Response = get_response_from_url(url)
        bs_page_with_matches: BeautifulSoup = BeautifulSoup(page_with_matches.content, 'lxml')

        list_of_links_to_matches += [HOST + path.get('href')
                                     for path in bs_page_with_matches.select('.table-main__tt>a')]
        print(f'Получены ссылки с матчами по дате {date_start}.')
        date_start += datetime.timedelta(1)
    else:
        print('Закончили получать ссылки на матч.')

    return list_of_links_to_matches


def get_data_about_match(url):
    """
    This function get content by url from match page and distributes
    information to the keys specified in the make_pattern_xlsx.py module
    in the column variable.

    :param url(url_type) -> link on match page.
    :return data_about_match -> data about the match.
    """
    match_page: Response = get_response_from_url(url)
    odds_table: Response = get_response_from_url(f'https://www.betexplorer.com/match-odds/'
                                                 f'{match_page.url.split("/")[-2]}'
                                                 f'/1/1x2/',
                                                 headers={'Referer': match_page.url})

    bs_match_page: BeautifulSoup = BeautifulSoup(match_page.content, 'lxml')
    bs_odds_table: BeautifulSoup = BeautifulSoup(odds_table.text.replace(r'\n', '\n').replace(r'\ '[0], ""),
                                                 'html.parser')

    data_about_match: dict[column_name, Any] = {col: '-' for col in make_pattern_xlsx.COLUMNS}

    # Set date and time
    date_and_time: list[numeric_str] = bs_match_page.select('#match-date')[0].get('data-dt').split(',')
    data_about_match['Дата'] = '.'.join(date_and_time[:3])
    data_about_match['Время'] = ':'.join(date_and_time[3:])

    # Set match name
    match_name: name = bs_match_page.select('.list-breadcrumb__item__in')[-1].get_text()
    data_about_match['Название матча'] = (match_name, match_page.url)

    # Set league name
    league_name: name = (f"{bs_match_page.select('.list-breadcrumb__item__in')[-3].get_text()}: "
                         f"{bs_match_page.select('.list-breadcrumb__item__in')[-2].get_text()}")
    data_about_match['Название лиги'] = (league_name, '/'.join(match_page.url.split('/')[:6]))

    # Set Game Scope
    try:
        game_scope: str = bs_match_page.select('#js-score')[0].get_text()
        data_about_match['Счет матча'] = game_scope
    except IndexError:
        data_about_match['Счет матча'] = '0:0'

    # Set the score for the first and second half
    try:
        first_half_score, second_half_score = bs_match_page.select('#js-partial')[0].get_text().split(', ')
        data_about_match['Счет первого тайма'] = first_half_score[1:]
        data_about_match['Счет второго тайма'] = second_half_score[:-1]
    except (ValueError, IndexError):
        data_about_match['Счет первого тайма'] = '0:0'
        data_about_match['Счет второго тайма'] = '0:0'

    # Set average odds
    average_odds: list[Tag] = bs_odds_table.select('#sortable-1>tfoot>tr>.table-main__detail-odds')
    if average_odds:
        data_about_match['Средний коэффициент на домашнюю команду'] = average_odds[0].get('data-odd')
        data_about_match['Средний коэффициент на ничью'] = average_odds[1].get('data-odd')
        data_about_match['Средний коэффициент на гостевую команду'] = average_odds[2].get('data-odd')

    # Set the minute at which the first goal was scored
    goals_tables: list[Tag] = [table for table in bs_match_page.select('.list-details--shooters>li>table')]
    if goals_tables:
        times_of_goals_scored: list[int] = []
        for table in goals_tables:
            for tr in table.select('tr'):
                cell_version_1 = tr.select('td')[0].get_text().replace('.', '')
                cell_version_2 = tr.select('td')[1].get_text().replace('.', '')

                if cell_version_2.isdigit() or '+' in cell_version_2:
                    if '+' in cell_version_2:
                        minutes: list[str] = cell_version_2.split('+')
                        if minutes[0].isdigit():
                            times_of_goals_scored.append(int(minutes[0]) + int(minutes[1]))
                    else:
                        times_of_goals_scored.append(int(cell_version_2))
                else:
                    if cell_version_1.isdigit():
                        times_of_goals_scored.append(int(cell_version_1))
                    else:
                        minutes: list[str] = cell_version_1.split('+')
                        if minutes[0].isdigit():
                            times_of_goals_scored.append(int(minutes[0]) + int(minutes[1]))
        if times_of_goals_scored:
            data_about_match['Минута, на которой был забит первый гол'] = min(times_of_goals_scored)
        else:
            data_about_match['Минута, на которой был забит первый гол'] = '-'
    else:
        data_about_match['Минута, на которой был забит первый гол'] = '-'

    # Set odds of winnings at bookmakers
    rows_of_bookmakers_with_odds: list[Tag] = bs_odds_table.select('#sortable-1>tbody>tr')
    for row in rows_of_bookmakers_with_odds:
        bookmaker_name: name = row.select('td>a.in-bookmaker-logo-link')[0].get_text()
        if bookmaker_name not in make_pattern_xlsx.COLUMNS[11:]:
            continue
        coefficient: float = row.select('td.table-main__detail-odds')[0].get('data-odd')
        data_about_match[bookmaker_name] = coefficient

    # Set maximal win coefficient
    list_of_odds_from_bookmakers: list[float] = [float(coefficient)
                                                 for coefficient in list(data_about_match.values())[11:]
                                                 if coefficient != '-']
    if list_of_odds_from_bookmakers:
        data_about_match['Максимальный коэффициент на победу фаворита'] = f'{max(list_of_odds_from_bookmakers):<04}'

    print(f'Спарсены данные о матче по ссылке {url}')

    return data_about_match


def write_data_about_match_in_xlsx_file(url: url_type) -> None:
    """Write data about match in Match_Statistic.xlsx with styles."""
    match_statistic_xlsx: Workbook = openpyxl.load_workbook('Match_Statistic.xlsx')
    match_statistic: Worksheet = match_statistic_xlsx['Match statistic']

    for link_on_match in get_list_of_links_to_matches(url):
        data = get_data_about_match(link_on_match)
        data['Название матча'], link_on_match = data['Название матча']
        data['Название лиги'], link_on_league = data['Название лиги']
        data_values = list(data.values())
        match_statistic.append(data_values)
        match_statistic[f'C{match_statistic.max_row}'].hyperlink = link_on_match
        match_statistic[f'D{match_statistic.max_row}'].hyperlink = link_on_league
        for coefficient in data_values[11:-1]:
            if coefficient != '-':
                if float(coefficient) == float(match_statistic[f'AG{match_statistic.max_row}'].value):
                    match_statistic[(f'{get_column_letter(list(data.values()).index(coefficient) + 1)}'
                                     f'{match_statistic.max_row}')].fill = PatternFill('solid',
                                                                                       bgColor=Color('FFFF00'),
                                                                                       fgColor=Color('FFFF00'))
        match_statistic_xlsx.save('Match_Statistic.xlsx')
        print(f'Данные о матче по ссылке {link_on_match} записаны в Match_Statistic.xlsx файл')


def run():
    if os.path.exists('Match_Statistic.xlsx'):
        write_data_about_match_in_xlsx_file(HOST + '/results/soccer/')
    else:
        make_pattern_xlsx.create_pattern_xlsx()
        write_data_about_match_in_xlsx_file(HOST + '/results/soccer/')


if __name__ == '__main__':
    run()