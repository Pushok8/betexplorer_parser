"""This module has one function to create a template xlsx and a constant COLUMNS, which contains the column name"""
import openpyxl
from openpyxl.styles.colors import Color
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.cell import get_column_letter
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font

from annotations import column_name

COLUMNS: list[column_name] = [
    'Дата',
    'Время',
    'Название матча',
    'Название лиги',
    'Счет матча',
    'Счет первого тайма',
    'Счет второго тайма',
    'Средний коэффициент на домашнюю команду',
    'Средний коэффициент на ничью',
    'Средний коэффициент на гостевую команду',
    'Минута, на которой был забит первый гол',
    '10Bet',
    '188BET',
    '1xBet',
    '888sport',
    'bet-at-home',
    'bet365',
    'Betclic',
    'Betfair',
    'Betsafe',
    'Betsson',
    'BetVictor',
    'Betway',
    'bwin',
    'ComeOn',
    'Interwetten',
    'Pinnacle',
    'SBOBET',
    'Unibet',
    'William Hill',
    'youwin',
    'Betfair Exchange',
    'Максимальный коэффициент на победу фаворита'
]


def create_pattern_xlsx():
    """This function to create template xlsx file with style. This file is called Match_Statistic.xlsx."""
    bold_side: Side = Side('hair', color=Color())
    bold_font: Font = Font('Roboto', bold=True)
    bold_border: Border = Border(left=bold_side, right=bold_side, top=bold_side, bottom=bold_side)
    column_number: int = 1

    match_statistic_xlsx: Workbook = openpyxl.Workbook()

    match_statistic_xlsx.remove(match_statistic_xlsx['Sheet'])
    match_statistic: Worksheet = match_statistic_xlsx.create_sheet('Match statistic')

    match_statistic.merge_cells('H1:J2')
    match_statistic.merge_cells('L1:AF2')

    match_statistic['H1'].font = bold_font
    match_statistic['L1'].font = bold_font
    match_statistic['H1'].alignment = Alignment('center')
    match_statistic['L1'].alignment = Alignment('center')
    match_statistic['H1'].border = bold_border
    match_statistic['L1'].border = bold_border
    match_statistic['H1'].value = 'Average odds'
    match_statistic['L1'].value = 'Названия букмекерских контор и коэффициенты на победу фаворита'

    for col_name in COLUMNS:
        match_statistic[f'{get_column_letter(column_number)}3'].font = bold_font
        match_statistic[f'{get_column_letter(column_number)}3'].border = bold_border
        match_statistic[f'{get_column_letter(column_number)}3'] = col_name
        column_number += 1



    match_statistic_xlsx.save('Match_Statistic.xlsx')
    match_statistic_xlsx.close()

